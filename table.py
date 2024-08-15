from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from config import TABLE_NAME
from pathlib import Path
from typing import Tuple, List


class FinancialTableManager:
    """
    Менеджер для работы с таблицей
    """

    TABLE_COLUMN_NAMES = ["Название", "Доход", "Расход", "Прибыль", "КПН"]
    COMPANY_FIXTURES = [
        ["Рога и копыта", 6000000, 800000, 0, 15],
        ["Ромашки", 5000000, 700000, 0, 20],
        ["Моя оборона", 4000000, 600000, 0, 25],
    ]

    def __init__(self, table_name: str):
        self.table_path = Path(table_name)

    def table_exists(self) -> bool:
        """Проверяет, существует ли таблица"""
        return self.table_path.exists()

    def _open_workbook(self) -> Tuple[Workbook, Worksheet]:
        """Открывает и возвращает активную книгу и лист."""
        wb = load_workbook(self.table_path, data_only=True)
        ws = wb.active
        return wb, ws

    def get_column_letter_from_idx(self, name: str) -> str:
        """Получает букву колонки"""
        wb, ws = self._open_workbook()
        for col in ws.iter_cols():
            if col[0].value == name:
                wb.close()
                return col[0].column_letter
        wb.close()
        raise ValueError(f"Column '{name}' not found in the table.")

    def create_table_with_data(self):
        """Заполняет таблицу данными, применяет стиль"""
        wb = Workbook()
        ws = wb.active
        ws.append(self.TABLE_COLUMN_NAMES)
        tab = Table(displayName="Table1", ref="A1:E5")
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(self.table_path)
        wb.close()
        self._set_table_fixtures()
        self._set_formula_for_profit_column()

    def _set_table_fixtures(self):
        """Внутренний метод заполняет таблицу данными из фикстур"""
        wb, ws = self._open_workbook()
        for row in self.COMPANY_FIXTURES:
            ws.append(row)
        wb.save(self.table_path)
        wb.close()

    def _set_formula_for_profit_column(self):
        """Внутренний метод, заполняет колонку "Прибыль" формулой"""
        column_idx = self.get_column_letter_from_idx("Прибыль")
        wb, ws = self._open_workbook()
        for row in range(2, ws.max_row + 1):
            profit_cell = ws[f"{column_idx}{row}"]
            if profit_cell.value is None or profit_cell.value == 0:
                ws[f"{column_idx}{row}"] = f"=(B{row}-C{row})*(1 - (E{row}/100))"
        wb.save(self.table_path)
        wb.close()

    def get_company_names(self) -> List[str]:
        """Возвращает список названий компаний из таблицы."""
        wb, ws = self._open_workbook()
        column_idx = self.get_column_letter_from_idx("Название")
        companies = [cell.value for cell in ws[column_idx][1:]]
        wb.close()
        return companies

    def get_company_data(self, company: str, data_type: str) -> float:
        """Возвращает значение данных для выбранной компании и типа данных."""
        wb, ws = self._open_workbook()
        column_idx = self.get_column_letter_from_idx(data_type)
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
            if row[0].value == company:
                data = row[ws[column_idx + "1"].column - 1].value
                wb.close()
                return data
        wb.close()
        raise ValueError(f"Данные для компании '{company}' не были найдены.")


if __name__ == "__main__":
    manager = FinancialTableManager(TABLE_NAME)
    manager.create_table_with_data()
